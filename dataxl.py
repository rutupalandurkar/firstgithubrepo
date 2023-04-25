import  openpyxl
wb=openpyxl.load_workbook('D:\\riya\\riya.xlsx')
Sheet=wb.active
r=Sheet.max_row
c=Sheet.max_column
print(r)
print(c)
for i in range(r+1):
    for j in range(c+1):
        print(Sheet.cell(row=i+1,column=j+1).value,end=' ')
    print()
for i in range(r+1,r+6):
    for j in range(c+1):
        Sheet.cell(row=i,column=j+1).value="hello"
wb.save('D:\\riya\\riya.xlsx')